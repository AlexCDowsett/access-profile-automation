"""
Access Profile Automation V2

A robust, senior-level implementation for parsing Excel access profile data.
Uses pandas for reliable data handling and includes comprehensive error handling.

Key improvements over V1:
- Uses pandas for robust Excel parsing
- Handles variable header rows automatically
- Clear separation of concerns with dedicated methods
- Comprehensive error handling and validation
- Type hints for better code clarity
- Detailed logging for debugging
"""

import os
import json
import logging
from typing import Dict, List, Tuple, Optional
from collections import defaultdict
import pandas as pd


class AccessProfileParser:
    """Handles parsing of Excel access profile data with robust error handling."""
    
    def __init__(self, file_path: str):
        """
        Initialize the parser with the Excel file path.
        
        Args:
            file_path: Path to the Excel file
            
        Raises:
            ValueError: If file_path is empty or invalid
        """
        if not file_path or not isinstance(file_path, str):
            raise ValueError("File path must be a non-empty string")
            
        self.file_path = os.path.abspath(file_path)
        self.logger = self._setup_logger()
        
        
    def _setup_logger(self) -> logging.Logger:
        """Set up logging for debugging and monitoring."""
        try:
            logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
            logger.setLevel(logging.INFO)
            
            if not logger.handlers:
                handler = logging.StreamHandler()
                formatter = logging.Formatter(
                    '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
                )
                handler.setFormatter(formatter)
                logger.addHandler(handler)
                
            return logger
        except Exception as e:
            # Fallback to basic logging if setup fails
            logging.basicConfig(level=logging.INFO)
            return logging.getLogger(__name__)
    
    def parse_excel(self) -> Tuple[List[str], List[str], pd.DataFrame]:
        """
        Parse the Excel file and extract categories, headings, and data.
        
        Returns:
            Tuple of (categories, headings, data_dataframe)
            
        Raises:
            FileNotFoundError: If the Excel file doesn't exist
            ValueError: If the Excel file structure is invalid
        """
        # Validate file exists and is accessible
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
            
        if not os.access(self.file_path, os.R_OK):
            raise PermissionError(f"Cannot read Excel file: {self.file_path}")
            
        # Check file extension
        if not self.file_path.lower().endswith(('.xlsx', '.xls')):
            raise ValueError(f"File must be an Excel file (.xlsx or .xls): {self.file_path}")
            
        try:
            # Read the entire Excel file without headers
            try:
                df = pd.read_excel(self.file_path, header=None, engine='openpyxl')
            except Exception as e:
                # Try with different engine if openpyxl fails
                try:
                    df = pd.read_excel(self.file_path, header=None, engine='xlrd')
                except Exception:
                    raise ValueError(f"Failed to read Excel file with any engine: {str(e)}")
                    
            if df.empty:
                raise ValueError("Excel file is empty or contains no data")
                
            self.logger.info(f"Successfully loaded Excel file with shape: {df.shape}")
            
            # Find the structure rows dynamically
            categories_row, headings_row, data_start_row = self._find_structure_rows(df)
            
            # Extract categories (row 0)
            categories = self._extract_categories(df, categories_row)
            
            # Extract headings (row 1) 
            headings = self._extract_headings(df, headings_row)
            
            # Extract data starting from row 3
            data_df = self._extract_data(df, data_start_row)
            
            self.logger.info(f"Found {len(categories)} categories, {len(headings)} headings, {len(data_df)} data rows")
            
            return categories, headings, data_df
            
        except Exception as e:
            self.logger.error(f"Error parsing Excel file: {str(e)}")
            raise ValueError(f"Failed to parse Excel file: {str(e)}")
    
    def _find_structure_rows(self, df: pd.DataFrame) -> Tuple[int, int, int]:
        """
        Dynamically find the row indices for categories, headings, and data start.
        
        Args:
            df: The loaded DataFrame
            
        Returns:
            Tuple of (categories_row, headings_row, data_start_row)
        """
        # Look for the categories row (contains 'Conductor', 'storm Contact', etc.)
        category_keywords = ['Conductor', 'storm Contact', 'UC', 'DataManagement', 'Dial', 'Flow']
        
        categories_row = None
        try:
            for idx, row in df.iterrows():
                if idx > 10:  # Safety limit to prevent infinite loops
                    break
                row_values = [str(val).strip() for val in row.values if pd.notna(val)]
                if any(keyword in row_values for keyword in category_keywords):
                    categories_row = idx
                    break
        except Exception as e:
            raise ValueError(f"Error searching for categories row: {str(e)}")
                
        if categories_row is None:
            raise ValueError(f"Could not find categories row in Excel file. Expected keywords: {category_keywords}")
            
        # Look for the headings row (contains 'Actions', 'Announcements', etc.)
        heading_keywords = ['Actions', 'Announcements', 'Menus', 'Parameters', 'Services']
        
        headings_row = None
        try:
            for idx, row in df.iterrows():
                if idx <= categories_row:
                    continue
                if idx > categories_row + 10:  # Safety limit
                    break
                row_values = [str(val).strip() for val in row.values if pd.notna(val)]
                if any(keyword in row_values for keyword in heading_keywords):
                    headings_row = idx
                    break
        except Exception as e:
            raise ValueError(f"Error searching for headings row: {str(e)}")
                
        if headings_row is None:
            raise ValueError(f"Could not find headings row in Excel file. Expected keywords: {heading_keywords}")
            
        # Data starts after the column headers row (headings_row + 1)
        data_start_row = headings_row + 1
        
        self.logger.info(f"Structure found - Categories: row {categories_row}, Headings: row {headings_row}, Data starts: row {data_start_row}")
        
        return categories_row, headings_row, data_start_row
    
    def _extract_categories(self, df: pd.DataFrame, row_idx: int) -> List[str]:
        """Extract category names from the specified row, handling merged cells."""
        try:
            # Use openpyxl to properly handle merged cells
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            categories = []
            current_category = None
            
            # Go through each column and extract categories
            for col_idx in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row_idx + 1, column=col_idx)  # openpyxl is 1-indexed
                    value = cell.value
                    
                    if value is not None and str(value).strip():
                        current_category = str(value).strip()
                        categories.append(current_category)
                except Exception as e:
                    self.logger.warning(f"Error reading cell at row {row_idx + 1}, col {col_idx}: {str(e)}")
                    continue
            
            if not categories:
                raise ValueError(f"No categories found in row {row_idx + 1}")
                
            return categories
        except Exception as e:
            raise ValueError(f"Failed to extract categories from row {row_idx + 1}: {str(e)}")
    
    def _extract_headings(self, df: pd.DataFrame, row_idx: int) -> List[str]:
        """Extract heading names from the specified row."""
        try:
            if row_idx >= len(df):
                raise ValueError(f"Row index {row_idx} is out of bounds for DataFrame with {len(df)} rows")
                
            row = df.iloc[row_idx]
            headings = [str(val).strip() for val in row.values if pd.notna(val)]
            
            if not headings:
                raise ValueError(f"No headings found in row {row_idx}")
                
            return headings
        except Exception as e:
            raise ValueError(f"Failed to extract headings from row {row_idx}: {str(e)}")
    
    def _extract_data(self, df: pd.DataFrame, start_row: int) -> pd.DataFrame:
        """Extract data rows starting from the specified row."""
        try:
            if start_row + 1 >= len(df):
                raise ValueError(f"Data start row {start_row + 1} is out of bounds for DataFrame with {len(df)} rows")
                
            # Skip the column headers row and get actual data
            data_df = df.iloc[start_row + 1:].copy()
            
            if data_df.empty:
                raise ValueError("No data rows found after header rows")
                
            # Remove completely empty rows
            data_df = data_df.dropna(how='all')
            
            if data_df.empty:
                raise ValueError("All data rows are empty")
                
            # Reset index for cleaner data
            data_df = data_df.reset_index(drop=True)
            
            return data_df
        except Exception as e:
            raise ValueError(f"Failed to extract data starting from row {start_row + 1}: {str(e)}")


class CategoryMapper:
    """Maps headings to their corresponding categories based on Excel column positions."""
    
    def __init__(self, categories: List[str], headings: List[str], file_path: str):
        """
        Initialize the mapper with categories, headings, and file path.
        
        Args:
            categories: List of category names
            headings: List of heading names
            file_path: Path to Excel file for merged cell analysis
            
        Raises:
            ValueError: If categories, headings, or file_path are invalid
        """
        if not categories or not isinstance(categories, list):
            raise ValueError("Categories must be a non-empty list")
        if not headings or not isinstance(headings, list):
            raise ValueError("Headings must be a non-empty list")
        if not file_path or not isinstance(file_path, str):
            raise ValueError("File path must be a non-empty string")
            
        self.categories = categories
        self.headings = headings
        self.file_path = file_path
        
        try:
            self.heading_to_category = self._build_mapping()
        except Exception as e:
            raise ValueError(f"Failed to build category mapping: {str(e)}")
        
    def _build_mapping(self) -> Dict[str, str]:
        """
        Build mapping from headings to categories based on Excel column positions.
        
        Returns:
            Dictionary mapping heading names to category names
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            mapping = {}
            
            # Get the actual column positions of categories and headings
            categories_positions = self._get_category_positions(ws)
            headings_positions = self._get_heading_positions(ws)
            
            if not categories_positions:
                raise ValueError("No category positions found")
            if not headings_positions:
                raise ValueError("No heading positions found")
            
            # Map each heading to its corresponding category based on column position
            for heading, heading_col in headings_positions.items():
                try:
                    # Find which category this heading belongs to based on column position
                    category = self._find_category_for_column(heading_col, categories_positions)
                    mapping[heading] = category
                except Exception as e:
                    raise ValueError(f"Failed to map heading '{heading}' to category: {str(e)}")
                    
            return mapping
        except Exception as e:
            raise ValueError(f"Failed to build category mapping: {str(e)}")
    
    def _get_category_positions(self, ws) -> Dict[str, int]:
        """Get the column positions of each category."""
        try:
            categories_positions = {}
            row_idx = 1  # Categories are in row 1
            
            if ws.max_column is None:
                raise ValueError("Worksheet has no columns")
                
            for col_idx in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    
                    if value is not None and str(value).strip():
                        category = str(value).strip()
                        categories_positions[category] = col_idx
                except Exception as e:
                    # Skip problematic cells
                    continue
            
            return categories_positions
        except Exception as e:
            raise ValueError(f"Failed to get category positions: {str(e)}")
    
    def _get_heading_positions(self, ws) -> Dict[str, int]:
        """Get the column positions of each heading."""
        try:
            headings_positions = {}
            row_idx = 2  # Headings are in row 2
            
            if ws.max_column is None:
                raise ValueError("Worksheet has no columns")
                
            for col_idx in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    
                    if value is not None and str(value).strip():
                        heading = str(value).strip()
                        headings_positions[heading] = col_idx
                except Exception as e:
                    # Skip problematic cells
                    continue
            
            return headings_positions
        except Exception as e:
            raise ValueError(f"Failed to get heading positions: {str(e)}")
    
    def _find_category_for_column(self, heading_col: int, categories_positions: Dict[str, int]) -> str:
        """Find which category a heading column belongs to."""
        try:
            if not categories_positions:
                raise ValueError("No category positions provided")
                
            if heading_col < 1:
                raise ValueError(f"Invalid column position: {heading_col}")
                
            # Sort categories by their column positions
            sorted_categories = sorted(categories_positions.items(), key=lambda x: x[1])
            
            # Find the category whose position is <= heading_col
            for i, (category, col_pos) in enumerate(sorted_categories):
                if i + 1 < len(sorted_categories):
                    next_category, next_col_pos = sorted_categories[i + 1]
                    if col_pos <= heading_col < next_col_pos:
                        return category
                else:
                    # Last category - covers everything from its position onwards
                    if col_pos <= heading_col:
                        return category
            
            # Fallback to first category
            return sorted_categories[0][0] if sorted_categories else self.categories[0]
        except Exception as e:
            raise ValueError(f"Failed to find category for column {heading_col}: {str(e)}")
    
    def get_category_for_heading(self, heading: str) -> str:
        """Get the category for a given heading."""
        if not heading or not isinstance(heading, str):
            raise ValueError("Heading must be a non-empty string")
            
        try:
            return self.heading_to_category.get(heading, self.categories[-1] if self.categories else "Unknown")
        except Exception as e:
            raise ValueError(f"Failed to get category for heading '{heading}': {str(e)}")


class OpenAccessProfilesXLSXV2:
    """
    Enhanced version of OpenAccessProfilesXLSX with robust parsing and better structure.
    
    Key improvements:
    - Uses pandas for reliable Excel parsing
    - Handles variable header rows automatically
    - Clear separation of concerns
    - Comprehensive error handling
    - Type hints and documentation
    """
    
    def __init__(self, file: str = 'AccessProfilesTEAM.xlsx'):
        """
        Initialize the access profiles parser.
        
        Args:
            file: Name of the Excel file to parse
            
        Raises:
            ValueError: If file parameter is invalid
            FileNotFoundError: If Excel file is not found
            RuntimeError: If parsing fails
        """
        if not file or not isinstance(file, str):
            raise ValueError("File parameter must be a non-empty string")
            
        try:
            # Set up file path
            base_dir = os.path.dirname(os.path.abspath(__file__))
            self.file_path = os.path.join(base_dir, file)
            
            # Initialize data structures
            self.access_profile_dict = self._create_nested_dict()
            self.filter = {}
            self.progress = 0
            
            # Parse the Excel file
            self._parse_file()
        except Exception as e:
            raise RuntimeError(f"Failed to initialize AccessProfilesXLSXV2: {str(e)}")
    
    def _create_nested_dict(self) -> Dict[str, Dict[str, Dict[str, List[str]]]]:
        """Create a 3-level nested defaultdict structure."""
        return defaultdict(lambda: defaultdict(dict))
    
    def _parse_file(self) -> None:
        """Parse the Excel file and populate the data structures."""
        try:
            # Parse the Excel file
            parser = AccessProfileParser(self.file_path)
            categories, headings, data_df = parser.parse_excel()
            
            if not categories:
                raise ValueError("No categories found in Excel file")
            if not headings:
                raise ValueError("No headings found in Excel file")
            if data_df.empty:
                raise ValueError("No data rows found in Excel file")
            
            # Create category mapper
            mapper = CategoryMapper(categories, headings, self.file_path)
            
            # Process each data row
            self._process_data_rows(data_df, headings, mapper)
            
            if not self.access_profile_dict:
                raise ValueError("No access profiles were successfully parsed")
                
        except Exception as e:
            raise RuntimeError(f"Failed to parse access profiles file: {str(e)}")
    
    def _process_data_rows(self, data_df: pd.DataFrame, headings: List[str], mapper: CategoryMapper) -> None:
        """
        Process each row of data and populate the access profile dictionary.
        Uses the exact same logic as V1 for data extraction.
        
        Args:
            data_df: DataFrame containing the data rows
            headings: List of heading names
            mapper: CategoryMapper instance for mapping headings to categories
        """
        # Use openpyxl to match V1's exact data extraction logic
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            if ws.max_row is None or ws.max_row < 3:
                raise ValueError("Worksheet has insufficient rows")
                
            # Find the data start row
            data_start_row = None
            for row_idx in range(1, min(ws.max_row + 1, 100)):  # Safety limit
                try:
                    cell = ws.cell(row=row_idx, column=3)  # Column C (Access Profile Name)
                    if cell.value is not None and str(cell.value).strip().startswith('TEAM -'):
                        data_start_row = row_idx
                        break
                except Exception:
                    continue
                    
            if data_start_row is None:
                raise ValueError("Could not find data start row (looking for 'TEAM -' prefix)")
        except Exception as e:
            raise ValueError(f"Failed to initialize Excel processing: {str(e)}")
        
        # Process each data row using V1's exact logic
        processed_count = 0
        error_count = 0
        
        for row_idx in range(data_start_row, ws.max_row + 1):
            try:
                # Extract row data exactly like V1 does
                row = ws[row_idx]
                row_no_none = [cell.value for cell in row if cell.value is not None]
                
                if len(row_no_none) < 2:
                    continue  # Skip rows without at least name and filter
                
                # Extract profile name and filter (same as V1)
                name = str(row_no_none[0]).strip()
                filter_val = str(row_no_none[1]).strip()
                
                if not name or not filter_val:
                    continue  # Skip rows with empty name or filter
                    
                row_no_none = row_no_none[2:]  # Skip name and filter
                
                # Store the filter
                self.filter[name] = filter_val
                
                # Process data using V1's exact logic
                cat_index = 0
                for i in range(len(headings)):
                    try:
                        if headings[i] in ['Agent Groups', 'Call Barring Profiles', 'Queries', 'Pacing Profiles', 'Flow Services']:
                            cat_index += 1
                        
                        if 2*i + 1 < len(row_no_none):
                            operator = str(row_no_none[2*i]).strip()
                            value = str(row_no_none[2*i + 1]).strip()
                            
                            # Only store non-empty, non-"none", non-"not in use" operators (same as V1)
                            if operator.lower() not in ['none', 'not in use']:
                                category = mapper.get_category_for_heading(headings[i])
                                self.access_profile_dict[name][category][headings[i]] = [operator, value]
                    except Exception as e:
                        # Skip this heading if there's an error
                        continue
                        
                processed_count += 1
                                
            except Exception as e:
                # Log the error but continue processing other rows
                error_count += 1
                if error_count <= 10:  # Limit error messages
                    print(f"Warning: Error processing row {row_idx}: {str(e)}")
                continue
                
        if processed_count == 0:
            raise ValueError("No data rows were successfully processed")
            
        if error_count > 0:
            print(f"Warning: {error_count} rows had processing errors")
    
    def _is_valid_operator(self, operator: str) -> bool:
        """
        Check if an operator should be stored (not empty, "none", or "not in use").
        This matches the V1 logic which checks the operator, not the value.
        
        Args:
            operator: The operator to check
            
        Returns:
            True if the operator should be stored, False otherwise
        """
        try:
            if not operator or not isinstance(operator, str):
                return False
                
            operator_lower = operator.lower().strip()
            invalid_values = ['none', 'not in use', 'n\\a', 'n/a', '']
            
            return operator_lower not in invalid_values
        except Exception:
            return False
    
    def print_progress_bar(self, current: Optional[int] = None, total: Optional[int] = None, 
                          bar_length: int = 20, prefix: str = 'Progress', suffix: str = '', 
                          show: bool = True) -> None:
        """Print a progress bar."""
        try:
            if not show:
                return
                
            if current is None:
                current = self.progress
                self.progress += 1
                
            if total is None:
                total = len(self.access_profile_dict.keys())
                
            if total == 0:
                return
                
            # Validate inputs
            if not isinstance(current, int) or current < 0:
                current = 0
            if not isinstance(total, int) or total <= 0:
                total = 1
            if not isinstance(bar_length, int) or bar_length <= 0:
                bar_length = 20
                
            # Ensure current doesn't exceed total
            current = min(current, total)
                
            percent = float(current) / total
            filled_length = int(bar_length * percent)
            bar = '\u2588' * filled_length + '-' * (bar_length - filled_length)
            
            print(f'\r{prefix} |{bar}| {percent*100:6.2f}% ({current}/{total}) {suffix}', end='')
            
            if current >= total:
                print()  # Newline on complete
        except Exception:
            # Silently fail if progress bar printing fails
            pass
    
    def to_json(self) -> str:
        """
        Convert the access profile dictionary to JSON string.
        
        Returns:
            JSON string representation of the data
            
        Raises:
            ValueError: If JSON serialization fails
        """
        try:
            if not self.access_profile_dict:
                return "{}"
                
            return json.dumps(self.access_profile_dict, indent=4, ensure_ascii=False)
        except Exception as e:
            raise ValueError(f"Failed to convert data to JSON: {str(e)}")
