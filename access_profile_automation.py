
import argparse

DEFAULT_FILENAME = 'B.X_CONDUCTOR_AZ Italy_AccessProfiles_V1.1_18092025 (1).xlsx'
FILENAME1 = 'B.X_CONDUCTOR_AZ Bene_AccessProfiles_V1_30092025.xlsx'
FILENAME2 = 'AccessProfilesTEAM.xlsx'
SHEETNAME = None

import openpyxl
from collections import defaultdict
import json


def main():
    parser = argparse.ArgumentParser(description="Process access profile Excel file.")
    parser.add_argument('--file', type=str, default=DEFAULT_FILENAME, help='Excel filename to process')
    args = parser.parse_args()

    f = OpenAccessProfilesXLSX(file=args.file, sheetname=SHEETNAME, debug=True)
    f.to_json(True)
    f.to_csv()


class OpenAccessProfilesXLSX():
    def __init__(self, file='AccessProfilesTEAM.xlsx', sheetname=None, debug=False):
        import os
        # Always use the absolute path relative to this script's directory
        self.filename = '.'.join(file.split('.')[:-1])
        base_dir = os.path.dirname(os.path.abspath(__file__))
        abs_file = os.path.join(base_dir, file)
        workbook = openpyxl.load_workbook(abs_file, data_only=True)

        if len(workbook.sheetnames) == 1:
            sheet = workbook.active
        elif sheetname is not None:
            try:
                sheet = workbook[sheetname]
            except KeyError:
                raise ValueError(f"Sheet '{sheetname}' not found in the workbook. Available sheets: {workbook.sheetnames}")
        else:
            for name in workbook.sheetnames:
                if 'accessprofiles' in name.lower():
                    print(f"Using sheet: {name}")
                    sheet = workbook[name]
                    break

                
        self.access_profile_dict = self._nested_dict()
        self.filter = {}
        self.progress = 0
        self.debug = debug
        self.categories = []
        current_row = 0
        for row in sheet.iter_rows(min_row=0, values_only=True):
            current_row += 1
            if any(cell in ['Conductor', 'storm Contact', 'UC', 'DataManagement', 'Dial ', 'Flow'] for cell in row):
                self.categories = [cell for cell in row if cell is not None]
                break
        
        if self.categories == []:
            raise ValueError("No categories found in the Excel sheet. Is correct sheet selected?")

        if self.debug:
            print("\nPRODUCTS/CATEGORY FOUND")
            print("=" * 50)
            print(self.categories)

        for row in sheet.iter_rows(min_row=current_row, values_only=True):
            current_row += 1
            if any(cell in ['Actions', 'Announcements', 'Menus'] for cell in row):
                self.headings = [cell for cell in row if cell is not None]
                #print(headings)
                #print(current_row)
                break
        
        if self.debug:
            print("\nCOLUMNS/TYPES FOUND")
            print("=" * 50)
            print(self.headings)

        for row in sheet.iter_rows(min_row=current_row, values_only=True):
            current_row += 1

            offset = 0
            for cell in row:
                if (cell != None and cell != ''):
                    if 'name' in cell.lower():
                        break
                offset += 1
            break


        for row in sheet.iter_rows(min_row=current_row, values_only=True):

            current_row += 1
            row = row[offset:]
            name = row[0]
            filter = row[1]
            row = row[2:]
            if name == None or name == '':
                break

            cat_index = 0
            for i in range(len(self.headings)):
                if self.headings[i] in ['Agent Groups', 'Call Barring Profiles', 'Queries', 'Pacing Profiles', 'Flow Services']:
                    cat_index += 1
                if row[2*i].lower() not in ['none', 'not in use']:
                        #print(name, filter, self.categories[cat_index], self.headings[i], row[2*i], row[2*i+1])
                    self.access_profile_dict[name][filter][self.categories[cat_index]][self.headings[i]] = [row[2*i], row[2*i+1]]
    
        if self.debug:
            sample_profile = list(self.access_profile_dict.keys())[0]
            sample_filter = list(self.access_profile_dict[sample_profile].keys())[0]
            sample_data = self.access_profile_dict[sample_profile][sample_filter]
            #print ("Offset to 'Name' column:", offset) 
            print("\nSAMPLE DATA FOUND")
            print("=" * 50)
            print(f"   Sample Profile: '{sample_profile}'")
            print(f"   Filter: '{sample_filter}'")
            print(f"   Category: '{list(sample_data.keys())[0]}'")
            
            # Show first category data
            if sample_data:
                first_category = list(sample_data.keys())[0]
                first_heading = list(sample_data[first_category].keys())[0]
                first_values = sample_data[first_category][first_heading]
                print(f"\n   Sample data: dict['{sample_profile}']['{sample_filter}']['{first_category}']['{first_heading}'] = ['{first_values[0]}','{first_values[1]}']")
    
    def _nested_dict(self):
        """Creates a 3-level nested defaultdict."""
        return defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    
    def print_progress_bar(self, current=None, total=None, bar_length=40, prefix='', suffix='', show=True):
        if current == None:
            current = self.progress
            self.progress += 1
        if total == None:
            total = len(self.access_profile_dict.keys())
        if not show:
            return
        percent = float(current) / total if total else 0
        filled_length = int(bar_length * percent)
        bar = '\u2588' * filled_length + '-' * (bar_length - filled_length)
        print(f'\r{prefix} |{bar}| {percent*100:6.2f}% ({current}/{total}) {suffix}', end='')
        if current >= total:
            print()  # Newline on complete 

    def to_json(self, write=False):
        import os
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
        os.makedirs(output_dir, exist_ok=True)
        if write:
            json_path = os.path.join(output_dir, self.filename + '.json')
            with open(json_path, 'w') as f:
                json.dump(self.access_profile_dict, f, indent=4)
            print(f"\n✅ JSON file '{json_path}' written successfully.")
        return json.dumps(self.access_profile_dict, indent=4)
        
    def to_dict(self, dict=None):
        if dict is None:
            dict = self.access_profile_dict
        return dict
    
    def to_csv(self):
        import csv
        import os
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
        os.makedirs(output_dir, exist_ok=True)
        csv_path = os.path.join(output_dir, self.filename + '.csv')
        with open(csv_path, mode='w', newline='') as file:
            writer = csv.writer(file)

            csv = [['Name', 'Filter', 'Product', 'Type', 'Operator', 'Value']]
            for name, filter_dict in self.access_profile_dict.items():
                for filter, category_dict in filter_dict.items():
                    for category, heading_dict in category_dict.items():
                        for heading, [operator, value] in heading_dict.items():
                            csv.append([name, filter, category, heading, operator, value])
            for csv_row in csv:
                writer.writerow(csv_row)
        print(f"\n✅ CSV file '{csv_path}' written successfully.")


if __name__ == "__main__":
    main()