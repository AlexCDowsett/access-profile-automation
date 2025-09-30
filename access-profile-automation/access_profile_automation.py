FILENAME = 'B.X_CONDUCTOR_AZ Italy_AccessProfiles_V1.1_18092025 (1).xlsx'
FILENAME = 'AccessProfilesTEAM.xlsx'
#FILENAME = 'B.X_CONDUCTOR_AZ Bene_AccessProfiles_V1_30092025.xlsx'
SHEETNAME = None

import openpyxl
from collections import defaultdict
import json


def main():
    f = OpenAccessProfilesXLSX(file=FILENAME, sheetname=SHEETNAME, debug=True)
    f.to_json(True)
    #f.to_csv()


class OpenAccessProfilesXLSX():
    def __init__(self, file='AccessProfilesTEAM.xlsx', sheetname=None, debug=False):
        import os
        # Always use the absolute path relative to this script's directory
        self.filename = '.'.join(file.split('.')[:-1])
        base_dir = os.path.dirname(os.path.abspath(__file__))
        abs_file = os.path.join(base_dir, file)
        workbook = openpyxl.load_workbook(abs_file, data_only=True)

        if len(workbook.sheetnames) == 1:
            sheet = workbook.active
        elif sheetname is not None:
            try:
                sheet = workbook[sheetname]
            except KeyError:
                raise ValueError(f"Sheet '{sheetname}' not found in the workbook. Available sheets: {workbook.sheetnames}")
        else:
            for name in workbook.sheetnames:
                if 'accessprofiles' in name.lower():
                    print(f"Using sheet: {name}")
                    sheet = workbook[name]
                    break

                
        self.access_profile_dict = self._nested_dict()
        self.filter = {}
        self.progress = 0
        self.debug = debug
        self.categories = []
        current_row = 0
        for row in sheet.iter_rows(min_row=0, values_only=True):
            current_row += 1
            if any(cell in ['Conductor', 'storm Contact', 'UC', 'DataManagement', 'Dial ', 'Flow'] for cell in row):
                self.categories = [cell for cell in row if cell is not None]
                break
        
        if self.categories == []:
            raise ValueError("No categories found in the Excel sheet. Is correct sheet selected?")

        if self.debug:
            print("\nCATEGORIES FOUND")
            print("=" * 50)
            print(self.categories)

        for row in sheet.iter_rows(min_row=current_row, values_only=True):
            current_row += 1
            if any(cell in ['Actions', 'Announcements', 'Menus'] for cell in row):
                self.headings = [cell for cell in row if cell is not None]
                #print(headings)
                #print(current_row)
                break
        
        if self.debug:
            print("\nHEADINGS/COLUMNS FOUND")
            print("=" * 50)
            print(self.headings)

        for row in sheet.iter_rows(min_row=current_row, values_only=True):
            current_row += 1
            row_no_none = [cell for cell in row if cell is not None]
            break

        for row in sheet.iter_rows(min_row=current_row, values_only=True):
            current_row += 1
            row_no_none = [cell for cell in row if cell is not None]
            name = row_no_none[0]
            filter = row_no_none[1]
            row_no_none = row_no_none[2:]

            cat_index = 0
            for i in range(len(self.headings)):
                if self.headings[i] in ['Agent Groups', 'Call Barring Profiles', 'Queries', 'Pacing Profiles', 'Flow Services']:
                    cat_index += 1
                if row_no_none[2*i].lower() not in ['none', 'not in use']:
                    self.access_profile_dict[name][filter][self.categories[cat_index]][self.headings[i]] = [row_no_none[2*i], row_no_none[2*i+1]]
    
        if self.debug:
            sample_profile = list(self.access_profile_dict.keys())[0]
            sample_filter = list(self.access_profile_dict[sample_profile].keys())[0]
            sample_data = self.access_profile_dict[sample_profile][sample_filter]
            
            print("\nSAMPLE DATA FOUND")
            print("=" * 50)
            print(f"   Sample Profile: '{sample_profile}'")
            print(f"   Filter: '{sample_filter}'")
            print(f"   Category: '{list(sample_data.keys())[0]}'")
            
            # Show first category data
            if sample_data:
                first_category = list(sample_data.keys())[0]
                first_heading = list(sample_data[first_category].keys())[0]
                first_values = sample_data[first_category][first_heading]
                print(f"   Sample data: dict['{sample_profile}']['{sample_filter}']['{first_category}']['{first_heading}'] = ['{first_values[0]}','{first_values[1]}']")
    
    def _nested_dict(self):
        """Creates a 3-level nested defaultdict."""
        return defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    
    def print_progress_bar(self, current=None, total=None, bar_length=40, prefix='Progress', suffix='', show=True):
        if current == None:
            current = self.progress
            self.progress += 1
        if total == None:
            total = len(self.access_profile_dict.keys())
        if not show:
            return
        percent = float(current) / total if total else 0
        filled_length = int(bar_length * percent)
        bar = '\u2588' * filled_length + '-' * (bar_length - filled_length)
        print(f'\r{prefix} |{bar}| {percent*100:6.2f}% ({current}/{total}) {suffix}', end='')
        if current >= total:
            print()  # Newline on complete 

    def to_json(self, write=False):
        if write:
            with open(self.filename + '.json', 'w') as f:
                json.dump(self.access_profile_dict, f, indent=4)
        return json.dumps(self.access_profile_dict, indent=4)
        
    def to_dict(self):
        return self.access_profile_dict
    
    def to_csv(self):
        import csv
        with open(self.filename + '.csv', mode='w', newline='') as file:
            writer = csv.writer(file)
            csv = [['Name']]

            for name, category in self.access_profile_dict.items():
                for category, headings in category.items():
                    csv.append([name])
                    for [heading, operatorvalue] in headings.items():
                        if heading not in csv[0]:
                            csv[0].append(heading)
                            csv[0].append("?")
                        print(operatorvalue)
                        csv[-1].append(operatorvalue[0])
                        csv[-1].append(operatorvalue[1])
                        
            for csv_row in csv:
                writer.writerow(csv_row)


if __name__ == "__main__":
    main()