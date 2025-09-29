import openpyxl
from collections import defaultdict
import time

import json

#dict['name']['UC']['User Groups'] = value

def placeholder(name, key, dict):
    time.sleep(0.03)


def main():
    f = OpenAccessProfilesXLSX()
    total = len(f.access_profile_dict.keys())
    for i, key in enumerate(f.access_profile_dict.keys(), 1):
        f.print_progress_bar(current=i, total=total, prefix='Processing Access Profiles:', suffix=key)
        placeholder(key, f.access_profile_dict[key]['filter'], f.access_profile_dict[key])  # Simulate work being done
    print()  # Newline after the last progress bar update
    print(f.to_json())


class OpenAccessProfilesXLSX():
    def __init__(self, file='AccessProfilesTEAM.xlsx'):
        import os
        # Always use the absolute path relative to this script's directory
        base_dir = os.path.dirname(os.path.abspath(__file__))
        abs_file = os.path.join(base_dir, file)
        workbook = openpyxl.load_workbook(abs_file, data_only=True)
        sheet = workbook.active
                
        self.access_profile_dict = self._nested_dict()

        self.progress = 0
        
        current_row = 0
        for row in sheet.iter_rows(min_row=0, max_row=1, values_only=True):
            current_row += 1
            if any(cell in ['Conductor', 'storm Contact', 'UC', 'DataManagement', 'Dial ', 'Flow'] for cell in row):
                categories = [cell for cell in row if cell is not None]
                #print(categories)
                #print(current_row)
                break

        for row in sheet.iter_rows(min_row=current_row, values_only=True):
            current_row += 1
            if any(cell in ['Actions', 'Announcements', 'Menus'] for cell in row):
                headings = [cell for cell in row if cell is not None]
                #print(headings)
                #print(current_row)
                break

        for row in sheet.iter_rows(min_row=current_row, values_only=True):
            current_row += 1
            row_no_none = [cell for cell in row if cell is not None]
            #print(row_no_none)
    
            #print(current_row)
            break

        for row in sheet.iter_rows(min_row=current_row, values_only=True):
            current_row += 1
            row_no_none = [cell for cell in row if cell is not None]
            name = row_no_none[0]
            filter = row_no_none[1]
            row_no_none = row_no_none[2:]

            cat_index = 0
            for i in range(len(headings)):
                if headings[i] in ['Agent Groups', 'Call Barring Profiles', 'Queries', 'Pacing Profiles', 'Flow Services']:
                    cat_index += 1
                self.access_profile_dict[name]['filter'] = filter
                if row_no_none[2*i].lower() not in ['none', 'not in use']:
                    self.access_profile_dict[name][categories[cat_index]][headings[i]] = [row_no_none[2*i], row_no_none[2*i+1]]
    
    def _nested_dict(self):
        """Creates a 3-level nested defaultdict."""
        return defaultdict(lambda: defaultdict(dict))
    
    def print_progress_bar(self, current=None, total=None, bar_length=40, prefix='Progress', suffix='', show=True):
        if current == None:
            current = self.progress
            self.progress += 1
        if total == None:
            total = len(self.access_profile_dict.keys())
        if not show:
            return
        percent = float(current) / total if total else 0
        filled_length = int(bar_length * percent)
        bar = '\u2588' * filled_length + '-' * (bar_length - filled_length)
        print(f'\r{prefix} |{bar}| {percent*100:6.2f}% ({current}/{total}) {suffix}', end='')
        if current >= total:
            print()  # Newline on complete 

    def to_json(self):
        return json.dumps(self.access_profile_dict, indent=4)
            

    


if __name__ == "__main__":
    main()